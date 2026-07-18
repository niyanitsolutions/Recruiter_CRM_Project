"""HRM — Holiday Service"""
import csv
import io
import re
import logging
from datetime import datetime, date, timezone, timedelta
from typing import Optional, List
from bson import ObjectId

from app.models.company.hrm_holiday import HolidayType

logger = logging.getLogger(__name__)

# ── Import parsing helpers ─────────────────────────────────────────────────────
# Canonical field ← accepted header aliases. Headers are compared after
# normalization: lowercased, trimmed, and runs of spaces/underscores collapsed
# to a single space — so "Holiday Name", "holiday name", "HOLIDAY NAME" and
# "Holiday_Name" all match. Unmapped columns are ignored.
_HEADER_ALIASES = {
    "name":         {"name", "holiday name", "holiday", "title", "holiday title"},
    "date":         {"date", "holiday date", "day", "on"},
    "holiday_type": {"type", "holiday type", "category", "holiday category"},
    "description":  {"description", "desc", "notes", "note", "remark", "remarks", "details"},
    "is_paid":      {"is paid", "paid", "paid holiday", "paid leave"},
    "is_recurring": {"is recurring", "recurring", "recurring every year",
                     "recurring every", "repeat", "yearly", "annual"},
}
_ALIAS_TO_FIELD = {alias: field for field, aliases in _HEADER_ALIASES.items() for alias in aliases}

# Date formats that carry a year vs. year-less (day+month only → resolved to the
# import's target year). Day-first (Indian) ordering is preferred over month-first.
_DATE_FORMATS_WITH_YEAR = [
    "%Y-%m-%d", "%Y/%m/%d",
    "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y",
    "%d-%b-%Y", "%d %b %Y", "%d-%B-%Y", "%d %B %Y",
    "%b %d %Y", "%B %d %Y", "%b %d, %Y", "%B %d, %Y",
    "%m/%d/%Y", "%m-%d-%Y",
]
_DATE_FORMATS_NO_YEAR = [
    "%d-%b", "%d %b", "%d-%B", "%d %B", "%d/%b", "%b-%d", "%b %d",
    "%d-%m", "%d/%m",
]


def _norm_header(h) -> str:
    return re.sub(r"[\s_]+", " ", str(h if h is not None else "").strip().lower())


def _parse_bool(val, default: bool) -> bool:
    """Yes/No, True/False, 1/0, Y/N (any case). Blank → default."""
    if val is None:
        return default
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    if not s:
        return default
    if s in ("1", "true", "yes", "y", "t", "paid", "recurring"):
        return True
    if s in ("0", "false", "no", "n", "f", "unpaid", "none"):
        return False
    return default


def _parse_holiday_type(val) -> HolidayType:
    """"National Holiday" → national, etc. Trims the redundant ' Holiday'
    suffix and lowercases. Unknown/blank → NATIONAL (never skips a row)."""
    s = re.sub(r"\s+", " ", str(val if val is not None else "").strip().lower())
    if s.endswith(" holiday"):
        s = s[: -len(" holiday")].strip()
    try:
        return HolidayType(s)
    except ValueError:
        return HolidayType.NATIONAL


def _parse_date(val, default_year: int) -> str:
    """Normalize to YYYY-MM-DD. Accepts native datetime/date cells (Excel),
    Excel serial numbers, and a wide range of string formats — with or without
    a year (a year-less date resolves to default_year)."""
    if val is None or (isinstance(val, str) and not val.strip()):
        raise ValueError("empty date")
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    # Excel serial date (1900 date system; day 0 == 1899-12-30). Guard against bool.
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=int(val))).strftime("%Y-%m-%d")
        except (ValueError, OverflowError):
            raise ValueError(f"invalid serial date '{val}'")
    s = str(val).strip()
    for fmt in _DATE_FORMATS_WITH_YEAR:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    for fmt in _DATE_FORMATS_NO_YEAR:
        try:
            return datetime.strptime(s, fmt).replace(year=default_year).strftime("%Y-%m-%d")
        except ValueError:
            continue
    raise ValueError(f"invalid date '{s}'")


class HolidayService:
    COL = "hrm_holidays"
    AUDIT_COL = "hrm_audit_logs"

    def __init__(self, db):
        self.db = db
        self.col = db[self.COL]

    @staticmethod
    def _serialize(doc: dict) -> dict:
        if not doc:
            return {}
        doc["id"] = str(doc.pop("_id", ""))
        for f in ("created_at", "updated_at"):
            val = doc.get(f)
            if isinstance(val, datetime):
                doc[f] = val.strftime("%Y-%m-%dT%H:%M:%S") + "Z"
        return doc

    async def _audit(self, action: str, entity_id: str, user_id: str,
                     company_id: str, changes: Optional[dict] = None, ip: Optional[str] = None):
        now = datetime.now(timezone.utc).replace(tzinfo=None)
        await self.db[self.AUDIT_COL].insert_one({
            "_id": str(ObjectId()),
            "company_id": company_id,
            "module": "holidays",
            "action": action,
            "entity_type": "holiday",
            "entity_id": entity_id,
            "user_id": user_id,
            "ip_address": ip,
            "changes": changes or {},
            "timestamp": now,
        })

    # ── CRUD ─────────────────────────────────────────────────────────────────

    async def create(self, data: dict, company_id: str, created_by: str,
                     ip: Optional[str] = None) -> dict:
        now = datetime.now(timezone.utc).replace(tzinfo=None)
        # Duplicate check: same company + same date
        existing = await self.col.find_one({
            "company_id": company_id,
            "date": data["date"],
            "is_deleted": False,
        })
        if existing:
            raise ValueError(f"A holiday already exists on {data['date']}")

        doc_id = str(ObjectId())
        doc = {
            "_id": doc_id,
            "company_id": company_id,
            "name": data["name"],
            "date": data["date"],
            "holiday_type": data.get("holiday_type", HolidayType.NATIONAL),
            "description": data.get("description"),
            "is_paid": data.get("is_paid", True),
            "is_recurring": data.get("is_recurring", False),
            "applicable_departments": data.get("applicable_departments", []),
            "applicable_locations": data.get("applicable_locations", []),
            "is_active": True,
            "created_by": created_by,
            "updated_by": None,
            "created_at": now,
            "updated_at": now,
            "is_deleted": False,
        }
        await self.col.insert_one(doc)
        await self._audit("holiday_created", doc_id, created_by, company_id,
                          {"name": doc["name"], "date": doc["date"]}, ip)
        return self._serialize(doc)

    async def list(
        self,
        company_id: str,
        year: Optional[int] = None,
        holiday_type: Optional[str] = None,
        department: Optional[str] = None,
        page: int = 1,
        page_size: int = 50,
    ) -> dict:
        query: dict = {"company_id": company_id, "is_deleted": False}
        if year:
            query["date"] = {"$gte": f"{year}-01-01", "$lte": f"{year}-12-31"}
        if holiday_type:
            query["holiday_type"] = holiday_type
        if department:
            query["$or"] = [
                {"applicable_departments": []},
                {"applicable_departments": department},
            ]
        total = await self.col.count_documents(query)
        skip = (page - 1) * page_size
        cursor = self.col.find(query).sort("date", 1).skip(skip).limit(page_size)
        items = [self._serialize(d) async for d in cursor]
        return {"items": items, "total": total, "page": page, "page_size": page_size}

    async def get(self, holiday_id: str, company_id: str) -> Optional[dict]:
        doc = await self.col.find_one({"_id": holiday_id, "company_id": company_id, "is_deleted": False})
        return self._serialize(doc) if doc else None

    async def update(self, holiday_id: str, data: dict, company_id: str,
                     updated_by: str, ip: Optional[str] = None) -> Optional[dict]:
        now = datetime.now(timezone.utc).replace(tzinfo=None)
        existing = await self.col.find_one({"_id": holiday_id, "company_id": company_id, "is_deleted": False})
        if not existing:
            return None
        # Duplicate date check (exclude self)
        if "date" in data and data["date"] != existing["date"]:
            dup = await self.col.find_one({
                "company_id": company_id,
                "date": data["date"],
                "is_deleted": False,
                "_id": {"$ne": holiday_id},
            })
            if dup:
                raise ValueError(f"Another holiday already exists on {data['date']}")

        changes_before = {k: existing.get(k) for k in data}
        update_set = {**{k: v for k, v in data.items() if v is not None},
                      "updated_by": updated_by, "updated_at": now}
        await self.col.update_one({"_id": holiday_id}, {"$set": update_set})
        await self._audit("holiday_updated", holiday_id, updated_by, company_id,
                          {"before": changes_before, "after": data}, ip)
        existing.update(update_set)
        return self._serialize(existing)

    async def delete(self, holiday_id: str, company_id: str, deleted_by: str,
                     ip: Optional[str] = None) -> bool:
        now = datetime.now(timezone.utc).replace(tzinfo=None)
        result = await self.col.update_one(
            {"_id": holiday_id, "company_id": company_id, "is_deleted": False},
            {"$set": {"is_deleted": True, "updated_at": now, "updated_by": deleted_by}},
        )
        if result.modified_count:
            await self._audit("holiday_deleted", holiday_id, deleted_by, company_id, ip=ip)
        return result.modified_count > 0

    # ── Utility ───────────────────────────────────────────────────────────────

    async def is_holiday(self, check_date: str, company_id: str,
                         department: Optional[str] = None) -> Optional[dict]:
        """Return holiday doc if date is a holiday for company/department, else None."""
        query: dict = {
            "company_id": company_id,
            "date": check_date,
            "is_active": True,
            "is_deleted": False,
        }
        if department:
            query["$or"] = [
                {"applicable_departments": []},
                {"applicable_departments": department},
            ]
        doc = await self.col.find_one(query)
        return self._serialize(doc) if doc else None

    async def get_holidays_in_range(self, company_id: str, from_date: str,
                                    to_date: str, department: Optional[str] = None) -> List[str]:
        """Return list of holiday date strings in range for overlap detection."""
        query: dict = {
            "company_id": company_id,
            "date": {"$gte": from_date, "$lte": to_date},
            "is_active": True,
            "is_deleted": False,
        }
        if department:
            query["$or"] = [
                {"applicable_departments": []},
                {"applicable_departments": department},
            ]
        cursor = self.col.find(query, {"date": 1})
        return [doc["date"] async for doc in cursor]

    # ── Import / Export ───────────────────────────────────────────────────────

    # ── Row extraction per file type ──────────────────────────────────────────

    @staticmethod
    def _rows_from_csv(text: str) -> List[dict]:
        reader = csv.DictReader(io.StringIO(text.strip()))
        return [dict(raw) for raw in reader]

    @staticmethod
    def _rows_from_xlsx(content: bytes) -> List[dict]:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            ws = wb.active
            headers, rows = None, []
            for r in ws.iter_rows(values_only=True):
                if r is None or all(c is None for c in r):
                    continue
                if headers is None:
                    headers = [str(c) if c is not None else "" for c in r]
                    continue
                rows.append({headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))})
            return rows
        finally:
            wb.close()

    @staticmethod
    def _rows_from_xls(content: bytes) -> List[dict]:
        import xlrd
        book = xlrd.open_workbook(file_contents=content)
        sheet = book.sheet_by_index(0)
        if sheet.nrows == 0:
            return []
        headers = [str(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
        rows = []
        for r in range(1, sheet.nrows):
            row = {}
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                val = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    val = xlrd.xldate_as_datetime(val, book.datemode)
                row[headers[c]] = val
            rows.append(row)
        return rows

    def _map_row(self, raw: dict) -> dict:
        """Map arbitrary source headers to canonical fields via _ALIAS_TO_FIELD."""
        mapped: dict = {}
        for k, v in raw.items():
            field = _ALIAS_TO_FIELD.get(_norm_header(k))
            if field and (field not in mapped or mapped[field] in (None, "")):
                mapped[field] = v
        return mapped

    # ── Public import entry points ────────────────────────────────────────────

    async def import_from_file(self, content: bytes, filename: str, company_id: str,
                               created_by: str, year: Optional[int] = None,
                               ip: Optional[str] = None) -> dict:
        """Import holidays from a CSV, XLSX or XLS file (auto-detected by extension)."""
        fn = (filename or "").lower()
        if fn.endswith(".xlsx"):
            raw_rows = self._rows_from_xlsx(content)
        elif fn.endswith(".xls"):
            raw_rows = self._rows_from_xls(content)
        else:  # .csv / .txt / unknown → treat as CSV text
            try:
                text = content.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = content.decode("latin-1")
            raw_rows = self._rows_from_csv(text)
        return await self._import_rows(raw_rows, company_id, created_by, year, ip)

    async def import_from_csv(self, csv_content: str, company_id: str,
                              created_by: str, ip: Optional[str] = None) -> dict:
        """Backward-compatible CSV entry point (older callers / tests)."""
        return await self._import_rows(self._rows_from_csv(csv_content), company_id, created_by, None, ip)

    async def _import_rows(self, raw_rows: List[dict], company_id: str, created_by: str,
                           year: Optional[int], ip: Optional[str]) -> dict:
        """Validate + normalize rows and batch-insert the valid, non-duplicate
        ones. Returns detailed per-row results.

        Uniqueness rule (unchanged from create()): one holiday per date per
        company — a normalized date already carries the year, so this is the
        same-tenant-same-year rule. Existing dates are pre-loaded once so a
        1000-row import issues a single read + a single insert_many, not N of each.
        """
        default_year = year or datetime.now().year

        existing_dates = set()
        async for d in self.col.find({"company_id": company_id, "is_deleted": False}, {"date": 1}):
            if d.get("date"):
                existing_dates.add(d["date"])

        created = skipped = 0
        errors: List[str] = []
        docs: List[dict] = []
        seen_in_batch = set()
        now = datetime.now(timezone.utc).replace(tzinfo=None)

        for i, raw in enumerate(raw_rows, start=2):   # row 1 = header
            mapped = self._map_row(raw)
            name = str(mapped.get("name") or "").strip()
            date_raw = mapped.get("date")

            if not name:
                errors.append(f"Row {i}: Missing Holiday Name"); skipped += 1
                logger.info("Holiday import skip (row %d): missing name", i); continue
            if date_raw is None or (isinstance(date_raw, str) and not date_raw.strip()):
                errors.append(f"Row {i}: Missing Date"); skipped += 1
                logger.info("Holiday import skip (row %d): missing date", i); continue
            try:
                iso_date = _parse_date(date_raw, default_year)
            except ValueError:
                errors.append(f"Row {i}: Invalid date ({str(date_raw).strip()})"); skipped += 1
                logger.info("Holiday import skip (row %d): invalid date %r", i, date_raw); continue

            if iso_date in existing_dates or iso_date in seen_in_batch:
                errors.append(f"Row {i}: Duplicate holiday — one already exists on {iso_date}")
                skipped += 1
                logger.info("Holiday import skip (row %d): duplicate on %s", i, iso_date); continue

            seen_in_batch.add(iso_date)
            desc = mapped.get("description")
            docs.append({
                "_id": str(ObjectId()),
                "company_id": company_id,
                "name": name,
                "date": iso_date,
                "holiday_type": _parse_holiday_type(mapped.get("holiday_type")).value,
                "description": (str(desc).strip() or None) if desc not in (None, "") else None,
                "is_paid": _parse_bool(mapped.get("is_paid"), True),
                "is_recurring": _parse_bool(mapped.get("is_recurring"), False),
                "applicable_departments": [],
                "applicable_locations": [],
                "is_active": True,
                "created_by": created_by,
                "updated_by": None,
                "created_at": now,
                "updated_at": now,
                "is_deleted": False,
            })
            created += 1

        if docs:
            await self.col.insert_many(docs)          # batch insert (supports 1000+ rows)
            await self._audit("holiday_imported", "bulk", created_by, company_id,
                              {"created": created, "skipped": skipped}, ip)

        return {
            "created": created,
            "skipped": skipped,
            "created_count": created,
            "skipped_count": skipped,
            "total": len(raw_rows),
            "errors": errors,
        }

    def export_to_csv(self, holidays: List[dict]) -> str:
        """Convert list of holiday dicts to CSV string."""
        output = io.StringIO()
        fields = ["name", "date", "holiday_type", "description", "is_paid", "is_recurring"]
        writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for h in holidays:
            writer.writerow(h)
        return output.getvalue()

    # ── Copy to next year ─────────────────────────────────────────────────────

    async def copy_to_next_year(self, company_id: str, created_by: str,
                                 ip: Optional[str] = None) -> dict:
        """
        Duplicate all recurring (or all, if include_all=True) holidays to next year.
        Returns {"created": N, "skipped": N}
        """
        current_year = date.today().year
        next_year = current_year + 1
        cursor = self.col.find({
            "company_id": company_id,
            "is_deleted": False,
            "is_recurring": True,
        })
        created = skipped = 0
        async for h in cursor:
            old_date = h["date"]  # YYYY-MM-DD
            try:
                dt = datetime.strptime(old_date, "%Y-%m-%d")
                new_date = dt.replace(year=next_year).strftime("%Y-%m-%d")
            except ValueError:
                skipped += 1
                continue
            # Skip if already exists
            dup = await self.col.find_one({
                "company_id": company_id,
                "date": new_date,
                "is_deleted": False,
            })
            if dup:
                skipped += 1
                continue
            try:
                await self.create({
                    "name": h["name"],
                    "date": new_date,
                    "holiday_type": h.get("holiday_type", HolidayType.NATIONAL),
                    "description": h.get("description"),
                    "is_paid": h.get("is_paid", True),
                    "is_recurring": True,
                    "applicable_departments": h.get("applicable_departments", []),
                    "applicable_locations": h.get("applicable_locations", []),
                }, company_id, created_by, ip)
                created += 1
            except ValueError:
                skipped += 1
        return {"created": created, "skipped": skipped, "year": next_year}
