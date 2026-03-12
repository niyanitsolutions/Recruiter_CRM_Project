"""
Export Service - Phase 5
Handles data export to Excel, CSV, and PDF
"""
from datetime import datetime, date, timedelta
from typing import Optional, List, Dict, Any
from bson import ObjectId
import io
import json

from app.models.company.import_export import (
    ImportExportType, ExportStatus, ExportFormat,
    ExportJobResponse, ExportJobListResponse,
    CreateExportRequest,
    IMPORT_EXPORT_TYPE_DISPLAY
)


class ExportService:
    """Service for data export operations"""
    
    def __init__(self, db):
        self.db = db
        self.export_jobs = db.export_jobs
    
    # ============== Export Job Management ==============
    
    async def create_export(
        self,
        data: CreateExportRequest,
        company_id: str,
        user_id: str
    ) -> ExportJobResponse:
        """Create an export job"""
        # Generate export name if not provided
        export_name = data.export_name or f"{data.entity_type.value}_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        job = {
            "id": str(ObjectId()),
            "company_id": company_id,
            "entity_type": data.entity_type.value,
            "export_name": export_name,
            "format": data.format.value,
            "filters": data.filters or {},
            "columns": data.columns or [],
            "sort_by": data.sort_by,
            "sort_order": data.sort_order,
            "include_headers": True,
            "status": ExportStatus.PENDING.value,
            "total_records": 0,
            "processed_records": 0,
            "created_at": datetime.utcnow(),
            "created_by": user_id,
            "is_deleted": False
        }
        
        await self.export_jobs.insert_one(job)
        
        # Process export (in production, this would be async/background)
        await self._process_export(job["id"], company_id)
        
        # Get updated job
        updated_job = await self.export_jobs.find_one({"id": job["id"]})
        
        response = ExportJobResponse(**updated_job)
        response.entity_type_display = IMPORT_EXPORT_TYPE_DISPLAY.get(
            ImportExportType(updated_job["entity_type"]), updated_job["entity_type"]
        )
        response.status_display = updated_job["status"].replace("_", " ").title()
        
        return response
    
    async def get_export(
        self,
        export_id: str,
        company_id: str
    ) -> Optional[ExportJobResponse]:
        """Get export job by ID"""
        job = await self.export_jobs.find_one({
            "id": export_id,
            "company_id": company_id,
            "is_deleted": False
        })
        
        if job:
            response = ExportJobResponse(**job)
            response.entity_type_display = IMPORT_EXPORT_TYPE_DISPLAY.get(
                ImportExportType(job["entity_type"]), job["entity_type"]
            )
            response.status_display = job["status"].replace("_", " ").title()
            return response
        return None
    
    async def list_exports(
        self,
        company_id: str,
        page: int = 1,
        page_size: int = 20,
        entity_type: Optional[ImportExportType] = None,
        status: Optional[ExportStatus] = None
    ) -> ExportJobListResponse:
        """List export jobs"""
        query = {"company_id": company_id, "is_deleted": False}
        
        if entity_type:
            query["entity_type"] = entity_type.value
        if status:
            query["status"] = status.value
        
        total = await self.export_jobs.count_documents(query)
        skip = (page - 1) * page_size
        
        cursor = self.export_jobs.find(query).sort("created_at", -1).skip(skip).limit(page_size)
        
        items = []
        async for job in cursor:
            response = ExportJobResponse(**job)
            response.entity_type_display = IMPORT_EXPORT_TYPE_DISPLAY.get(
                ImportExportType(job["entity_type"]), job["entity_type"]
            )
            response.status_display = job["status"].replace("_", " ").title()
            items.append(response)
        
        return ExportJobListResponse(
            items=items,
            total=total,
            page=page,
            page_size=page_size
        )
    
    # ============== Export Processing ==============
    
    async def _process_export(self, export_id: str, company_id: str):
        """Process an export job"""
        job = await self.export_jobs.find_one({"id": export_id})
        if not job:
            return
        
        try:
            # Update status to processing
            await self.export_jobs.update_one(
                {"id": export_id},
                {"$set": {"status": ExportStatus.PROCESSING.value, "started_at": datetime.utcnow()}}
            )
            
            # Get data based on entity type
            entity_type = ImportExportType(job["entity_type"])
            data = await self._get_export_data(
                entity_type,
                company_id,
                job.get("filters", {}),
                job.get("columns", []),
                job.get("sort_by"),
                job.get("sort_order", "asc")
            )
            
            # Generate file based on format
            export_format = ExportFormat(job["format"])
            file_content, file_name = await self._generate_file(
                data,
                entity_type,
                export_format,
                job["export_name"]
            )
            
            # In production, upload to S3/cloud storage
            # For now, we'll simulate with a placeholder URL
            file_url = f"/exports/{company_id}/{file_name}"
            
            # Update job with results
            await self.export_jobs.update_one(
                {"id": export_id},
                {"$set": {
                    "status": ExportStatus.COMPLETED.value,
                    "completed_at": datetime.utcnow(),
                    "total_records": len(data),
                    "processed_records": len(data),
                    "file_url": file_url,
                    "file_name": file_name,
                    "file_size": len(file_content) if isinstance(file_content, bytes) else len(file_content.encode()),
                    "expires_at": datetime.utcnow() + timedelta(days=7)
                }}
            )
            
        except Exception as e:
            await self.export_jobs.update_one(
                {"id": export_id},
                {"$set": {
                    "status": ExportStatus.FAILED.value,
                    "completed_at": datetime.utcnow(),
                    "error_message": str(e)
                }}
            )
    
    async def _get_export_data(
        self,
        entity_type: ImportExportType,
        company_id: str,
        filters: Dict[str, Any],
        columns: List[str],
        sort_by: Optional[str],
        sort_order: str
    ) -> List[Dict[str, Any]]:
        """Get data for export"""
        # Build query
        query = {"company_id": company_id, "is_deleted": False}
        
        # Apply filters
        for key, value in filters.items():
            if value is not None:
                query[key] = value
        
        # Get collection based on entity type
        if entity_type == ImportExportType.CANDIDATES:
            collection = self.db.candidates
            default_columns = [
                "full_name", "email", "mobile", "current_company",
                "current_designation", "experience_years", "current_ctc",
                "expected_ctc", "notice_period", "location", "source", "status"
            ]
        elif entity_type == ImportExportType.CLIENTS:
            collection = self.db.clients
            default_columns = [
                "name", "industry", "website", "address", "city",
                "state", "contact_name", "contact_email", "contact_phone", "status"
            ]
        elif entity_type == ImportExportType.JOBS:
            collection = self.db.jobs
            default_columns = [
                "title", "client_name", "positions", "job_type",
                "work_mode", "location", "min_experience", "max_experience",
                "min_salary", "max_salary", "status"
            ]
        elif entity_type == ImportExportType.APPLICATIONS:
            collection = self.db.applications
            default_columns = [
                "candidate_name", "job_title", "client_name",
                "status", "applied_date", "current_stage"
            ]
        elif entity_type == ImportExportType.USERS:
            collection = self.db.users
            default_columns = [
                "full_name", "email", "mobile", "role",
                "department", "designation", "is_active"
            ]
        else:
            return []
        
        # Use specified columns or defaults
        selected_columns = columns if columns else default_columns
        
        # Build projection
        projection = {col: 1 for col in selected_columns}
        projection["_id"] = 0
        
        # Sort
        sort_direction = 1 if sort_order == "asc" else -1
        sort_field = sort_by or "created_at"
        
        # Execute query
        cursor = collection.find(query, projection).sort(sort_field, sort_direction)
        
        data = []
        async for doc in cursor:
            # Clean up data
            row = {}
            for col in selected_columns:
                value = doc.get(col)
                if isinstance(value, datetime):
                    value = value.isoformat()
                elif isinstance(value, date):
                    value = value.isoformat()
                elif isinstance(value, ObjectId):
                    value = str(value)
                row[col] = value
            data.append(row)
        
        return data
    
    async def _generate_file(
        self,
        data: List[Dict[str, Any]],
        entity_type: ImportExportType,
        export_format: ExportFormat,
        export_name: str
    ) -> tuple:
        """Generate export file"""
        if export_format == ExportFormat.CSV:
            return self._generate_csv(data, export_name)
        elif export_format == ExportFormat.EXCEL:
            return self._generate_excel(data, export_name)
        elif export_format == ExportFormat.JSON:
            return self._generate_json(data, export_name)
        else:
            raise ValueError(f"Unsupported export format: {export_format}")
    
    def _generate_csv(self, data: List[Dict[str, Any]], export_name: str) -> tuple:
        """Generate CSV file"""
        if not data:
            return "", f"{export_name}.csv"
        
        import csv
        
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
        
        content = output.getvalue()
        return content, f"{export_name}.csv"
    
    def _generate_excel(self, data: List[Dict[str, Any]], export_name: str) -> tuple:
        """Generate Excel file"""
        if not data:
            return b"", f"{export_name}.xlsx"
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Export"
            
            # Header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            
            # Write headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header.replace("_", " ").title())
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
            
            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            content = output.getvalue()
            
            return content, f"{export_name}.xlsx"
            
        except ImportError:
            # Fallback to CSV if openpyxl not available
            return self._generate_csv(data, export_name)
    
    def _generate_json(self, data: List[Dict[str, Any]], export_name: str) -> tuple:
        """Generate JSON file"""
        content = json.dumps(data, indent=2, default=str)
        return content, f"{export_name}.json"
    
    # ============== Quick Export Methods ==============
    
    async def export_candidates(
        self,
        company_id: str,
        filters: Optional[Dict[str, Any]] = None,
        format: ExportFormat = ExportFormat.EXCEL
    ) -> bytes:
        """Quick export candidates"""
        data = await self._get_export_data(
            ImportExportType.CANDIDATES,
            company_id,
            filters or {},
            [],
            "created_at",
            "desc"
        )
        
        content, _ = await self._generate_file(
            data,
            ImportExportType.CANDIDATES,
            format,
            "candidates_export"
        )
        
        return content
    
    async def export_report_data(
        self,
        report_data: List[Dict[str, Any]],
        report_name: str,
        format: ExportFormat = ExportFormat.EXCEL
    ) -> tuple:
        """Export report data to file"""
        return await self._generate_file(
            report_data,
            ImportExportType.CANDIDATES,  # Dummy, not used for report exports
            format,
            report_name.replace(" ", "_").lower()
        )