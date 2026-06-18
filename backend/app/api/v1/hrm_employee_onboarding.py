"""HRM — Employee Self-Onboarding Form Link & Employee Export"""

import os
import io
import csv
import uuid
import pathlib
import logging
from datetime import datetime, timezone, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, BackgroundTasks
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.core.dependencies import get_company_db, require_hrm_module, require_permissions

logger = logging.getLogger(__name__)

_BACKEND_ROOT = pathlib.Path(__file__).resolve().parent.parent.parent.parent

# Authenticated router (HR/Admin actions)
router = APIRouter(prefix="/hrm/employees", tags=["HRM - Employee Onboarding"])

# Public router (token-based, no auth)
public_router = APIRouter(prefix="/public", tags=["HRM - Employee Onboarding (Public)"])


# ── Request schema ─────────────────────────────────────────────────────────────

class OnboardingLinkRequest(BaseModel):
    email: Optional[str] = None
    frontend_base_url: Optional[str] = None


# ── Helper: resolve token across all company DBs ──────────────────────────────

async def _find_token(token: str):
    """Search all company databases for an employee onboarding token.
    Returns (token_doc, company_db, company_id) or raises 404."""
    from app.core.database import get_master_db as _master, get_company_db as _cdb
    master_db = _master()
    tenants = await master_db.tenants.find({}, {"company_id": 1}).to_list(length=500)
    for t in tenants:
        cid = t.get("company_id")
        if not cid:
            continue
        cdb = _cdb(cid)
        doc = await cdb.employee_onboarding_tokens.find_one({"_id": token})
        if doc:
            return doc, cdb, cid
    raise HTTPException(status_code=404, detail="Invalid or expired link.")


def _validate_token_expiry(token_doc: dict):
    """Raise 410 if the token is used or expired."""
    if token_doc.get("used"):
        raise HTTPException(status_code=410, detail="This link has already been used.")
    exp = token_doc["expires_at"]
    if exp.tzinfo is None:
        exp = exp.replace(tzinfo=timezone.utc)
    if exp < datetime.now(timezone.utc):
        raise HTTPException(status_code=410, detail="This link has expired.")


# ── Generate Onboarding Link (authenticated) ──────────────────────────────────

@router.post("/generate-onboarding-link")
async def generate_employee_onboarding_link(
    background_tasks: BackgroundTasks,
    body: OnboardingLinkRequest = OnboardingLinkRequest(),
    cu: dict = Depends(require_hrm_module),
    db=Depends(get_company_db),
    _perm=Depends(require_permissions(["hrm:employees:manage"])),
):
    """Generate a unique one-time employee self-onboarding URL (expires in 7 days)."""
    from app.core.config import settings as _settings

    token = uuid.uuid4().hex
    now = datetime.now(timezone.utc)
    await db.employee_onboarding_tokens.insert_one({
        "_id": token,
        "company_id": cu["company_id"],
        "employee_email": body.email,
        "created_by": cu["id"],
        "created_at": now,
        "expires_at": now + timedelta(days=7),
        "used": False,
        "used_at": None,
        "employee_id": None,
    })

    base = (body.frontend_base_url or "").rstrip("/")
    form_url = f"{base}/employee-onboard/{token}"

    email_sent = False
    if body.email:
        if _settings.EMAIL_ENABLED:
            try:
                from app.services.email_service import send_employee_onboarding_link_email
                background_tasks.add_task(
                    send_employee_onboarding_link_email,
                    to_email=body.email,
                    form_url=form_url,
                    sent_by_name=cu.get("full_name", "HR Team"),
                    company_name=cu.get("company_name", ""),
                    company_id=cu.get("company_id", ""),
                )
                email_sent = True
            except Exception as e:
                logger.warning("Could not queue onboarding email: %s", e)
        else:
            logger.info(
                "[EMPLOYEE ONBOARDING] Email not sent (EMAIL_ENABLED=False). to=%s token=%s",
                body.email, token,
            )

    return {
        "success": True,
        "token": token,
        "form_url": form_url,
        "email_sent": email_sent,
        "message": "Onboarding link generated" + (" and emailed" if email_sent else ""),
    }


# ── Export Employees (authenticated) ──────────────────────────────────────────

@router.get("/export")
async def export_employees(
    format: str = Query("csv", pattern="^(csv|xlsx|pdf)$"),
    status: Optional[str] = None,
    department_id: Optional[str] = None,
    search: Optional[str] = None,
    cu: dict = Depends(require_hrm_module),
    db=Depends(get_company_db),
    _perm=Depends(require_permissions(["hrm:employees:view"])),
):
    """Export the employee list as CSV, XLSX, or PDF (respects current filters)."""
    q: dict = {"company_id": cu["company_id"], "is_deleted": False}
    if status:
        q["employment_status"] = status
    if department_id:
        q["department_id"] = department_id
    if search:
        q["$or"] = [
            {"full_name":   {"$regex": search, "$options": "i"}},
            {"email":       {"$regex": search, "$options": "i"}},
            {"employee_id": {"$regex": search, "$options": "i"}},
        ]

    cursor = db.hrm_employees.find(q, {
        "_id": 1, "employee_id": 1, "full_name": 1, "email": 1, "phone": 1,
        "department_name": 1, "designation_name": 1, "employment_type": 1,
        "employment_status": 1, "date_of_joining": 1,
        "date_of_birth": 1, "gender": 1,
        "bank_details": 1, "address_info": 1, "emergency_contacts": 1,
        "emergency_contact": 1, "qualifications": 1, "background_check": 1,
        "documents": 1,
    }).sort("created_at", -1)
    employees = await cursor.to_list(length=None)

    HEADERS = [
        "Employee ID", "Employee Name", "Email", "Mobile Number",
        "Department", "Designation", "Employment Type", "Status",
        "Profile Completion %", "Joining Date",
    ]

    def _pct(emp):
        addr = emp.get("address_info") or {}
        bank = emp.get("bank_details") or {}
        sections = [
            bool(emp.get("phone") and emp.get("date_of_birth") and emp.get("gender")),
            bool((emp.get("department_id") or emp.get("department_name")) and emp.get("date_of_joining")),
            bool(bank.get("bank_name") and bank.get("account_number")),
            bool((emp.get("emergency_contacts") or []) or emp.get("emergency_contact")),
            bool(emp.get("qualifications")),
            bool((emp.get("background_check") or {}).get("status")),
            bool(emp.get("documents")),
        ]
        return f"{round(sum(sections) / len(sections) * 100)}%"

    def _doj(emp):
        v = emp.get("date_of_joining")
        if v is None:
            return ""
        if hasattr(v, "strftime"):
            return v.strftime("%Y-%m-%d")
        return str(v)[:10]

    def _row(emp):
        return [
            emp.get("employee_id", ""),
            emp.get("full_name", ""),
            emp.get("email", ""),
            emp.get("phone", ""),
            emp.get("department_name", ""),
            emp.get("designation_name", ""),
            (emp.get("employment_type") or "").replace("_", " ").title(),
            (emp.get("employment_status") or "").replace("_", " ").title(),
            _pct(emp),
            _doj(emp),
        ]

    now_str = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"employees_{now_str}"

    # ── CSV ────────────────────────────────────────────────────────────────────
    if format == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(HEADERS)
        for emp in employees:
            writer.writerow(_row(emp))
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
        )

    # ── XLSX ───────────────────────────────────────────────────────────────────
    elif format == "xlsx":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employees"

        header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, emp in enumerate(employees, 2):
            for col_idx, value in enumerate(_row(emp), 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-size columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            iter([buf.read()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
        )

    # ── PDF ────────────────────────────────────────────────────────────────────
    elif format == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=landscape(A4),
            rightMargin=10 * mm, leftMargin=10 * mm,
            topMargin=15 * mm, bottomMargin=15 * mm,
        )
        styles = getSampleStyleSheet()
        elements = [
            Paragraph("Employee Report", styles["Title"]),
            Spacer(1, 6 * mm),
        ]

        table_data = [HEADERS] + [_row(emp) for emp in employees]
        t = Table(table_data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#6366F1")),
            ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 7),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FF")]),
            ("GRID",          (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("ALIGN",         (0, 0), (-1, -1), "LEFT"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING",   (0, 0), (-1, -1), 4),
        ]))
        elements.append(t)
        doc.build(elements)
        buf.seek(0)
        return StreamingResponse(
            iter([buf.read()]),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}.pdf"'},
        )


# ── Public: Validate Token ─────────────────────────────────────────────────────

@public_router.get("/employee-onboarding/{token}")
async def validate_employee_onboarding_token(token: str):
    """Validate token — called by the public form on mount to confirm the link is live."""
    token_doc, _, company_id = await _find_token(token)
    _validate_token_expiry(token_doc)
    return {"success": True, "token": token, "company_id": company_id}


# ── Public: Submit Onboarding Form ─────────────────────────────────────────────

@public_router.post("/employee-onboarding/{token}")
async def submit_employee_onboarding(token: str, data: dict):
    """
    Accept employee self-onboarding data (personal info, emergency contact,
    bank details, qualifications) and create an HRM employee record.
    Employment status is set to 'pending_hr_review' for HR to complete.
    Token is marked used after successful creation.
    """
    token_doc, company_db, company_id = await _find_token(token)
    _validate_token_expiry(token_doc)

    now = datetime.now(timezone.utc)
    employee_uuid = str(uuid.uuid4())

    # Auto-generate sequential employee ID
    count = await company_db.hrm_employees.count_documents({"company_id": company_id})
    emp_number = f"EMP{(count + 1):04d}"

    # Emergency contact
    ec = data.get("emergency_contact") or {}
    emergency_contacts = []
    if ec.get("name") and ec.get("phone"):
        emergency_contacts = [{
            "name": ec.get("name", ""),
            "relationship": ec.get("relationship", ""),
            "phone": ec.get("phone", ""),
        }]

    # Bank details
    bank = data.get("bank_details") or {}
    bank_details = {
        "bank_name": bank.get("bank_name") or "",
        "account_number": bank.get("account_number") or "",
        "ifsc_code": bank.get("ifsc_code") or "",
        "account_holder_name": bank.get("account_holder_name") or "",
    } if any(v for v in bank.values() if v) else None

    # Address
    addr = data.get("address_info") or {}
    address_info = {
        "street": addr.get("street") or data.get("current_address") or "",
        "city": addr.get("city") or "",
        "state": addr.get("state") or "",
        "zip_code": addr.get("zip_code") or "",
        "country": "India",
    } if (addr or data.get("current_address")) else None

    # Qualifications
    qualifications = []
    for q in (data.get("qualifications") or []):
        title = q.get("degree") or q.get("title") or ""
        if title:
            qualifications.append({
                "type": "academic",
                "title": title,
                "institution": q.get("institution") or q.get("college") or "",
                "year": int(q["year"]) if q.get("year") else None,
                "grade": q.get("grade") or "",
            })

    emp_doc = {
        "_id": employee_uuid,
        "company_id": company_id,
        "employee_id": emp_number,
        "full_name": (data.get("full_name") or "").strip(),
        "email": (data.get("email") or "").strip().lower(),
        "phone": (data.get("mobile") or data.get("phone") or "").strip(),
        "gender": data.get("gender") or None,
        "date_of_birth": data.get("date_of_birth") or None,
        "blood_group": data.get("blood_group") or None,
        "address": data.get("current_address") or None,
        "address_info": address_info,
        "permanent_address": data.get("permanent_address") or None,
        "emergency_contacts": emergency_contacts,
        "photo_url": None,
        "pan_number": data.get("pan_number") or None,
        "aadhaar_number": data.get("aadhaar_number") or None,
        # Employment info — HR completes these after review
        "department_id": None,
        "department_name": None,
        "designation_id": None,
        "designation_name": None,
        "reporting_manager_id": None,
        "reporting_manager_name": None,
        "employment_type": "full_time",
        "employment_status": "pending_hr_review",
        "date_of_joining": None,
        "work_location": None,
        "shift_start_time": "09:00",
        "shift_end_time": "18:00",
        "salary": {
            "ctc": 0, "basic": 0, "hra": 0, "special_allowance": 0,
            "pf_employee": 0, "pf_employer": 0, "professional_tax": 0,
            "gross_salary": 0, "net_salary": 0,
        },
        "bank_details": bank_details,
        "qualifications": qualifications,
        "documents": [],
        "background_check": None,
        "crm_user_id": None,
        "hrm_onboarding_id": None,
        "source": "onboarding_form",
        "employee_profile_status": "incomplete",
        "created_by": "self_onboarding",
        "created_at": now,
        "updated_at": now,
        "is_deleted": False,
        "deleted_at": None,
    }

    await company_db.hrm_employees.insert_one(emp_doc)

    # Mark token as used — store employee UUID for subsequent uploads
    await company_db.employee_onboarding_tokens.update_one(
        {"_id": token},
        {"$set": {"used": True, "used_at": now, "employee_id": employee_uuid}},
    )

    return {"success": True, "employee_id": employee_uuid, "employee_number": emp_number}


# ── Public: Upload Profile Photo ───────────────────────────────────────────────

@public_router.post("/employee-onboarding/{token}/photo")
async def upload_onboarding_photo(token: str, file: UploadFile = File(...)):
    """Upload profile photo after onboarding form has been submitted."""
    token_doc, company_db, _ = await _find_token(token)

    employee_id = token_doc.get("employee_id")
    if not employee_id:
        raise HTTPException(
            status_code=400,
            detail="Please submit the onboarding form before uploading a photo.",
        )

    ALLOWED = {".jpg", ".jpeg", ".png", ".webp"}
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in ALLOWED:
        raise HTTPException(status_code=400, detail="Only JPG, JPEG, PNG, WEBP images are allowed.")

    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Photo must be under 5 MB.")

    upload_dir = _BACKEND_ROOT / "uploads" / "hrm_docs"
    upload_dir.mkdir(parents=True, exist_ok=True)
    fname = f"{uuid.uuid4()}{ext}"
    (upload_dir / fname).write_bytes(contents)
    photo_url = f"/api/v1/uploads/hrm_docs/{fname}"

    await company_db.hrm_employees.update_one(
        {"_id": employee_id},
        {"$set": {"photo_url": photo_url, "updated_at": datetime.now(timezone.utc)}},
    )
    return {"success": True, "photo_url": photo_url}


# ── Public: Upload Document ─────────────────────────────────────────────────────

@public_router.post("/employee-onboarding/{token}/document")
async def upload_onboarding_document(
    token: str,
    doc_type: str = Query(
        ...,
        pattern="^(aadhaar|pan|degree_cert|resume|other|experience_letter|relieving_letter|payslip)$",
    ),
    file: UploadFile = File(...),
):
    """Upload a document (Aadhaar, PAN, Resume, Other) after form submission."""
    token_doc, company_db, _ = await _find_token(token)

    employee_id = token_doc.get("employee_id")
    if not employee_id:
        raise HTTPException(
            status_code=400,
            detail="Please submit the onboarding form before uploading documents.",
        )

    ALLOWED_EXT = {".pdf", ".jpg", ".jpeg", ".png", ".doc", ".docx"}
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in ALLOWED_EXT:
        raise HTTPException(
            status_code=400,
            detail="Only PDF, images, and Word documents are allowed.",
        )

    contents = await file.read()
    if len(contents) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Document must be under 10 MB.")

    upload_dir = _BACKEND_ROOT / "uploads" / "hrm_docs"
    upload_dir.mkdir(parents=True, exist_ok=True)
    fname = f"{uuid.uuid4()}{ext}"
    (upload_dir / fname).write_bytes(contents)
    file_url = f"/api/v1/uploads/hrm_docs/{fname}"

    doc_labels = {
        "aadhaar":           "Aadhaar Card",
        "pan":               "PAN Card",
        "degree_cert":       "Degree / Provisional Certificate",
        "resume":            "Resume",
        "other":             "Document",
        "experience_letter": "Experience Letter",
        "relieving_letter":  "Relieving Letter",
        "payslip":           "Latest Payslip",
    }
    doc_entry = {
        "doc_id": str(uuid.uuid4()),
        "doc_type": doc_type,
        "doc_name": doc_labels.get(doc_type, "Document"),
        "file_url": file_url,
        "original_filename": file.filename or fname,
        "status": "pending",
        "uploaded_by": "self",
        "uploaded_at": datetime.now(timezone.utc),
        "version": 1,
        "version_history": [],
    }

    await company_db.hrm_employees.update_one(
        {"_id": employee_id},
        {
            "$push": {"documents": doc_entry},
            "$set":  {"updated_at": datetime.now(timezone.utc)},
        },
    )
    return {"success": True, "file_url": file_url, "doc_type": doc_type}
